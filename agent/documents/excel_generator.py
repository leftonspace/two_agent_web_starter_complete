"""
PHASE 2.4: Excel Generator

Generate .xlsx Excel spreadsheets with multiple sheets, formatting, and formulas.

Features:
- Multi-sheet workbooks
- Header formatting (bold, colored backgrounds)
- Column width adjustment
- Freeze panes
- Formulas (SUM, AVERAGE, etc.)
- Cell alignment and styling
- Read data from existing Excel files

Usage:
    generator = ExcelGenerator()

    sheets = {
        "Sheet1": [
            ["Name", "Score", "Grade"],
            ["Alice", 95, "A"],
            ["Bob", 87, "B"]
        ]
    }

    generator.create_workbook(sheets, Path("output/grades.xlsx"))
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Generate Excel spreadsheets (.xlsx) with formatting and formulas.

    Supports:
    - Multiple sheets
    - Header formatting
    - Column widths
    - Freeze panes
    - Formulas
    - Reading existing files
    """

    def __init__(self):
        """Initialize Excel generator"""
        self.wb = None

    def create_workbook(
        self,
        sheets: Dict[str, List[List[Any]]],
        output_path: Path,
        formatting: Optional[Dict[str, Dict]] = None
    ) -> Path:
        """
        Create Excel workbook with multiple sheets.

        Args:
            sheets: Dictionary of sheet_name -> data
                   Data format: [["Header1", "Header2"], ["Value1", "Value2"], ...]
            output_path: Path to save workbook
            formatting: Optional formatting per sheet

        Returns:
            Path to created workbook

        Raises:
            ImportError: If openpyxl not installed
            Exception: If creation fails

        Example:
            sheets = {
                "Employees": [
                    ["Name", "Department", "Salary"],
                    ["Alice", "Engineering", 120000],
                    ["Bob", "Sales", 95000]
                ],
                "Summary": [
                    ["Total Employees", 2],
                    ["Average Salary", 107500]
                ]
            }

            formatting = {
                "Employees": {
                    "header_row": 1,
                    "freeze_panes": "A2",
                    "column_widths": {"A": 20, "B": 15, "C": 12}
                }
            }

            generator.create_workbook(sheets, Path("employees.xlsx"), formatting)
        """
        try:
            from openpyxl import Workbook

            output_path = Path(output_path)
            self.wb = Workbook()

            # Remove default sheet
            if self.wb.active:
                self.wb.remove(self.wb.active)

            # Create sheets
            for sheet_name, data in sheets.items():
                ws = self.wb.create_sheet(title=sheet_name)

                # Write data
                for row_idx, row_data in enumerate(data, start=1):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                # Apply formatting if provided
                if formatting and sheet_name in formatting:
                    self._apply_sheet_formatting(ws, formatting[sheet_name])

            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Save workbook
            self.wb.save(str(output_path))

            logger.info(f"Generated Excel workbook with {len(sheets)} sheet(s): {output_path}")
            return output_path

        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel generation. "
                "Install with: pip install openpyxl"
            )
        except Exception as e:
            logger.error(f"Failed to generate Excel workbook: {e}")
            raise

    def _apply_sheet_formatting(self, ws, fmt: Dict[str, Any]) -> None:
        """
        Apply formatting to worksheet.

        Args:
            ws: Worksheet object
            fmt: Formatting dictionary

        Format options:
            header_row: Row number for header (1-indexed)
            freeze_panes: Cell to freeze panes at (e.g., "A2")
            column_widths: Dict of column letter -> width (e.g., {"A": 20})
        """
        from openpyxl.styles import Font, PatternFill, Alignment

        # Header row styling
        if fmt.get("header_row"):
            header_row = fmt["header_row"]
            for cell in ws[header_row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="4472C4",
                    end_color="4472C4",
                    fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze panes
        if fmt.get("freeze_panes"):
            ws.freeze_panes = fmt["freeze_panes"]

        # Column widths
        if fmt.get("column_widths"):
            for col, width in fmt["column_widths"].items():
                ws.column_dimensions[col].width = width

        # Auto-filter
        if fmt.get("auto_filter"):
            ws.auto_filter.ref = fmt["auto_filter"]

    def add_formula(
        self,
        file_path: Path,
        sheet_name: str,
        cell: str,
        formula: str
    ) -> Optional[Any]:
        """
        Add formula to cell in existing file.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name
            cell: Cell reference (e.g., "D2")
            formula: Excel formula (e.g., "=SUM(A2:C2)")

        Returns:
            Calculated value (requires re-opening with data_only=True)

        Example:
            generator.add_formula(
                Path("grades.xlsx"),
                "Sheet1",
                "D2",
                "=SUM(B2:C2)"
            )
        """
        try:
            from openpyxl import load_workbook

            # Load workbook
            wb = load_workbook(str(file_path))
            ws = wb[sheet_name]

            # Add formula
            ws[cell] = formula

            # Save
            wb.save(str(file_path))

            logger.debug(f"Added formula to {sheet_name}!{cell}: {formula}")

            # Try to get calculated value
            try:
                wb_data = load_workbook(str(file_path), data_only=True)
                return wb_data[sheet_name][cell].value
            except:
                return None

        except Exception as e:
            logger.error(f"Failed to add formula: {e}")
            raise

    def read_data(
        self,
        file_path: Path,
        sheet_name: str,
        range_: Optional[str] = None
    ) -> List[List[Any]]:
        """
        Read data from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to read
            range_: Optional cell range (e.g., "A1:C10"), None reads all

        Returns:
            List of rows, each row is a list of cell values

        Example:
            data = generator.read_data(
                Path("data.xlsx"),
                "Sheet1",
                "A1:C10"
            )
        """
        try:
            from openpyxl import load_workbook

            # Load with data_only to get calculated values
            wb = load_workbook(str(file_path), data_only=True)
            ws = wb[sheet_name]

            data = []

            if range_:
                # Read specific range
                cells = ws[range_]
                for row in cells:
                    if not isinstance(row, tuple):
                        # Single cell
                        row = (row,)
                    row_data = [cell.value for cell in row]
                    data.append(row_data)
            else:
                # Read all rows
                for row in ws.iter_rows():
                    row_data = [cell.value for cell in row]
                    data.append(row_data)

            logger.debug(f"Read {len(data)} rows from {sheet_name}")
            return data

        except Exception as e:
            logger.error(f"Failed to read Excel data: {e}")
            raise

    def add_chart(
        self,
        file_path: Path,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        position: str = "E2"
    ) -> None:
        """
        Add chart to worksheet.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name
            chart_type: Chart type ("bar", "line", "pie")
            data_range: Data range for chart (e.g., "A1:C10")
            position: Position to place chart (e.g., "E2")

        Example:
            generator.add_chart(
                Path("sales.xlsx"),
                "Q1 Sales",
                "bar",
                "A1:B10",
                "D2"
            )
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.chart import BarChart, LineChart, PieChart, Reference

            wb = load_workbook(str(file_path))
            ws = wb[sheet_name]

            # Create chart based on type
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                raise ValueError(f"Unsupported chart type: {chart_type}")

            # Parse data range
            data = Reference(ws, range_string=data_range)
            chart.add_data(data, titles_from_data=True)

            # Add chart to sheet
            ws.add_chart(chart, position)

            # Save
            wb.save(str(file_path))

            logger.debug(f"Added {chart_type} chart to {sheet_name} at {position}")

        except Exception as e:
            logger.error(f"Failed to add chart: {e}")
            raise


# Convenience function
def generate_excel_workbook(
    sheets: Dict[str, List[List[Any]]],
    output_path: Path,
    formatting: Optional[Dict[str, Dict]] = None
) -> Path:
    """
    Convenience function to generate Excel workbook.

    Args:
        sheets: Dictionary of sheet_name -> data
        output_path: Output path
        formatting: Optional formatting

    Returns:
        Path to generated workbook

    Example:
        generate_excel_workbook(
            {"Sheet1": [["A", "B"], [1, 2]]},
            Path("output.xlsx")
        )
    """
    generator = ExcelGenerator()
    return generator.create_workbook(sheets, output_path, formatting)
